from __future__ import annotations

import math
import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

import networkx as nx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


SCHEMA_VERSION = 5
NODE_HEADERS = ("id", "label", "x", "y", "color")
EDGE_HEADERS = ("source", "target", "weight", "bold")
NODE_COLOR_PALETTE = (
    "#4C78A8",
    "#F58518",
    "#54A24B",
    "#E45756",
    "#72B7B2",
    "#B279A2",
    "#FF9DA6",
    "#9D755D",
)
DEFAULT_NODE_COLOR = NODE_COLOR_PALETTE[0]
COLOR_PATTERN = re.compile(r"^#[0-9A-Fa-f]{6}$")
MIN_NODE_COORDINATE = -1.0
MAX_NODE_COORDINATE = 2.0
MIN_VIEW_ZOOM = 0.1
MAX_VIEW_ZOOM = 3.0
MAX_VIEW_PAN = 10000.0


class GraphWorkbookError(ValueError):
    """Ошибка структуры или данных Excel-файла графа."""


def validate_weight(value: Any) -> float:
    try:
        weight = float(value)
    except (TypeError, ValueError) as exc:
        raise GraphWorkbookError("Вес связи должен быть числом от -1 до 1.") from exc
    if not math.isfinite(weight) or not -1.0 <= weight <= 1.0:
        raise GraphWorkbookError("Вес связи должен находиться в диапазоне от -1 до 1.")
    return weight


def validate_color(value: Any, default: str = DEFAULT_NODE_COLOR) -> str:
    color = default if value in (None, "") else str(value).strip()
    if not COLOR_PATTERN.fullmatch(color):
        raise GraphWorkbookError("Цвет вершины должен иметь формат #RRGGBB.")
    return color.upper()


def validate_bold(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "да", "1"}:
        return True
    if normalized in {"false", "no", "нет", "0"}:
        return False
    raise GraphWorkbookError("Признак жирной стрелки должен быть TRUE или FALSE.")


def rename_node(graph: nx.DiGraph, node_id: str, new_label: Any) -> None:
    if node_id not in graph:
        raise GraphWorkbookError("Выбранная вершина не найдена.")
    label = str(new_label).strip() if new_label is not None else ""
    if not label:
        raise GraphWorkbookError("Введите новое название вершины.")
    graph.nodes[node_id]["label"] = label


def add_connection(
    graph: nx.DiGraph,
    source: str,
    target: str,
    forward_weight: Any,
    *,
    bidirectional: bool = False,
    reverse_weight: Any | None = None,
    bold: bool = False,
    reverse_bold: bool = False,
) -> None:
    if source not in graph or target not in graph:
        raise GraphWorkbookError("Сначала создайте обе вершины связи.")
    if source == target:
        raise GraphWorkbookError("Связь вершины с самой собой пока не поддерживается.")
    graph.add_edge(source, target, weight=validate_weight(forward_weight), bold=validate_bold(bold))
    if bidirectional:
        reverse = forward_weight if reverse_weight is None else reverse_weight
        graph.add_edge(
            target,
            source,
            weight=validate_weight(reverse),
            bold=validate_bold(reverse_bold),
        )


def graph_to_excel_bytes(graph: nx.DiGraph) -> bytes:
    workbook = _build_workbook(graph)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def save_graph_to_excel(graph: nx.DiGraph, path: str | Path) -> Path:
    destination = Path(path).resolve()
    if destination.suffix.lower() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.stem}.tmp.xlsx")
    temporary.write_bytes(graph_to_excel_bytes(graph))
    os.replace(temporary, destination)
    return destination


def load_graph_from_excel(source: str | Path | BinaryIO) -> nx.DiGraph:
    if isinstance(source, (str, Path)):
        path = Path(source).resolve()
        if not path.exists():
            raise GraphWorkbookError(f"Файл не найден: {path}")
        workbook_source: str | Path | BinaryIO = path
    else:
        workbook_source = source

    try:
        workbook = load_workbook(workbook_source, data_only=True, read_only=True)
    except Exception as exc:
        raise GraphWorkbookError("Не удалось открыть Excel-файл.") from exc

    try:
        if "Вершины" not in workbook.sheetnames or "Связи" not in workbook.sheetnames:
            raise GraphWorkbookError("В книге должны быть листы «Вершины» и «Связи».")
        nodes_ws = workbook["Вершины"]
        edges_ws = workbook["Связи"]
        _require_headers(nodes_ws, NODE_HEADERS[:4])
        _require_headers(edges_ws, EDGE_HEADERS[:3])

        graph = nx.DiGraph()
        for row_number, values in enumerate(nodes_ws.iter_rows(min_row=2, values_only=True), start=2):
            node_id, label, x, y = values[:4]
            color_value = values[4] if len(values) > 4 else None
            default_color = NODE_COLOR_PALETTE[(row_number - 2) % len(NODE_COLOR_PALETTE)]
            if node_id in (None, ""):
                continue
            node_id = str(node_id)
            if node_id in graph:
                raise GraphWorkbookError(f"Повторяющийся id вершины в строке {row_number}: {node_id}")
            graph.add_node(
                node_id,
                label=str(label) if label not in (None, "") else node_id,
                x=_coordinate(x, row_number, "x"),
                y=_coordinate(y, row_number, "y"),
                color=validate_color(color_value, default=default_color),
            )

        for row_number, values in enumerate(edges_ws.iter_rows(min_row=2, values_only=True), start=2):
            source_id, target_id, weight = values[:3]
            bold_value = values[3] if len(values) > 3 else False
            if source_id in (None, "") and target_id in (None, "") and weight in (None, ""):
                continue
            source_id = str(source_id)
            target_id = str(target_id)
            if source_id not in graph or target_id not in graph:
                raise GraphWorkbookError(
                    f"Строка {row_number} листа «Связи» ссылается на отсутствующую вершину."
                )
            if source_id == target_id:
                raise GraphWorkbookError(f"Петля в строке {row_number} не поддерживается.")
            graph.add_edge(source_id, target_id, weight=validate_weight(weight), bold=validate_bold(bold_value))
        if "Настройки" in workbook.sheetnames:
            settings = {
                str(key): value
                for key, value, *_rest in workbook["Настройки"].iter_rows(min_row=2, values_only=True)
                if key not in (None, "")
            }
            _restore_view_settings(graph, settings)
        return graph
    finally:
        workbook.close()


def load_graph_from_excel_bytes(data: bytes) -> nx.DiGraph:
    return load_graph_from_excel(BytesIO(data))


def _build_workbook(graph: nx.DiGraph) -> Workbook:
    workbook = Workbook()
    nodes_ws = workbook.active
    nodes_ws.title = "Вершины"
    edges_ws = workbook.create_sheet("Связи")
    labeled_edges_ws = workbook.create_sheet("Связи по названиям")
    settings_ws = workbook.create_sheet("Настройки")

    nodes_ws.append(NODE_HEADERS)
    for node_id, attrs in graph.nodes(data=True):
        nodes_ws.append(
            [
                str(node_id),
                str(attrs.get("label", node_id)),
                _optional_float(attrs.get("x")),
                _optional_float(attrs.get("y")),
                validate_color(attrs.get("color", DEFAULT_NODE_COLOR)),
            ]
        )

    edges_ws.append(EDGE_HEADERS)
    labeled_edges_ws.append(EDGE_HEADERS)
    for source, target, attrs in graph.edges(data=True):
        weight = validate_weight(attrs.get("weight", 0.0))
        bold = validate_bold(attrs.get("bold", False))
        edges_ws.append([str(source), str(target), weight, bold])
        labeled_edges_ws.append(
            [
                str(graph.nodes[source].get("label", source)),
                str(graph.nodes[target].get("label", target)),
                weight,
                bold,
            ]
        )

    settings_ws.append(["key", "value"])
    settings_ws.append(["schema_version", SCHEMA_VERSION])
    settings_ws.append(["saved_at", datetime.now().isoformat(timespec="seconds")])
    settings_ws.append(["graph_type", "networkx.DiGraph"])
    settings_ws.append(["weight_range", "[-1, 1], zero is valid"])
    settings_ws.append(["view_zoom", _view_value(graph.graph.get("view_zoom"), 1.0, MIN_VIEW_ZOOM, MAX_VIEW_ZOOM)])
    settings_ws.append(["view_pan_x", _view_value(graph.graph.get("view_pan_x"), 0.0, -MAX_VIEW_PAN, MAX_VIEW_PAN)])
    settings_ws.append(["view_pan_y", _view_value(graph.graph.get("view_pan_y"), 0.0, -MAX_VIEW_PAN, MAX_VIEW_PAN)])

    _format_data_sheet(nodes_ws, widths=(18, 36, 14, 14, 14), table_name="NodesTable")
    _format_data_sheet(edges_ws, widths=(18, 18, 14, 14), table_name="EdgesTable")
    _format_data_sheet(
        labeled_edges_ws,
        widths=(36, 36, 14, 14),
        table_name="LabeledEdgesTable",
    )
    _format_data_sheet(settings_ws, widths=(22, 38), table_name="SettingsTable")
    for ws in (edges_ws, labeled_edges_ws):
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            row[0].number_format = "0.000"
    for ws in (nodes_ws, edges_ws, labeled_edges_ws, settings_ws):
        ws.sheet_view.showGridLines = False
    return workbook


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _view_value(value: Any, default: float, minimum: float, maximum: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(number):
        return default
    return max(minimum, min(maximum, number))


def _restore_view_settings(graph: nx.DiGraph, settings: dict[str, Any]) -> None:
    graph.graph["view_zoom"] = _view_value(
        settings.get("view_zoom"),
        1.0,
        MIN_VIEW_ZOOM,
        MAX_VIEW_ZOOM,
    )
    graph.graph["view_pan_x"] = _view_value(
        settings.get("view_pan_x"),
        0.0,
        -MAX_VIEW_PAN,
        MAX_VIEW_PAN,
    )
    graph.graph["view_pan_y"] = _view_value(
        settings.get("view_pan_y"),
        0.0,
        -MAX_VIEW_PAN,
        MAX_VIEW_PAN,
    )


def _coordinate(value: Any, row_number: int, name: str) -> float | None:
    if value in (None, ""):
        return None
    try:
        coordinate = float(value)
    except (TypeError, ValueError) as exc:
        raise GraphWorkbookError(f"Некорректная координата {name} в строке {row_number}.") from exc
    if not math.isfinite(coordinate):
        raise GraphWorkbookError(f"Некорректная координата {name} в строке {row_number}.")
    return max(MIN_NODE_COORDINATE, min(MAX_NODE_COORDINATE, coordinate))


def _require_headers(ws: Any, expected: tuple[str, ...]) -> None:
    actual = tuple(ws.cell(1, column).value for column in range(1, len(expected) + 1))
    if actual != expected:
        raise GraphWorkbookError(f"Лист «{ws.title}»: ожидались столбцы {', '.join(expected)}.")


def _format_data_sheet(ws: Any, widths: tuple[int, ...], table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    table = Table(displayName=table_name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
