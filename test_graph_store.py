from __future__ import annotations

import tempfile
import unittest
import math
from io import BytesIO
from pathlib import Path

import networkx as nx
from openpyxl import load_workbook

from app import (
    LAYOUT_HEIGHT,
    LAYOUT_WIDTH,
    MIN_NODE_CENTER_DISTANCE,
    MIN_NODE_EDGE_DISTANCE,
    apply_spring_layout,
    count_edge_crossings,
    separate_nodes_from_edges,
)
from graph_component import apply_position_updates
from graph_store import (
    GraphWorkbookError,
    add_connection,
    graph_to_excel_bytes,
    load_graph_from_excel,
    load_graph_from_excel_bytes,
    rename_node,
    save_graph_to_excel,
)


class GraphStoreTests(unittest.TestCase):
    def test_edge_crossing_counter_ignores_shared_endpoints_and_reciprocal_edges(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("A", x=0.0, y=0.0)
        graph.add_node("B", x=1.0, y=1.0)
        graph.add_node("C", x=0.0, y=1.0)
        graph.add_node("D", x=1.0, y=0.0)
        graph.add_edges_from((("A", "B"), ("B", "A"), ("C", "D"), ("A", "C")))

        self.assertEqual(count_edge_crossings(graph), 1)

    def test_rebuild_finds_crossing_free_layout_for_planar_complete_graph(self) -> None:
        graph = nx.DiGraph()
        graph.add_edges_from(
            (str(source), str(target))
            for source, target in nx.complete_graph(4).edges()
        )

        crossings = apply_spring_layout(graph, minimum_distance=120.0)

        self.assertEqual(crossings, 0)
        self.assertEqual(count_edge_crossings(graph), 0)

    def test_rebuild_moves_node_away_from_an_edge_and_its_label(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("A", x=0.1, y=0.5)
        graph.add_node("B", x=0.5, y=0.5)
        graph.add_node("C", x=0.9, y=0.5)
        graph.add_edge("A", "C")

        separate_nodes_from_edges(graph)

        distance = abs(graph.nodes["B"]["y"] - 0.5) * LAYOUT_HEIGHT
        self.assertGreaterEqual(distance, MIN_NODE_EDGE_DISTANCE - 0.5)

    def test_rebuild_separates_disconnected_nodes(self) -> None:
        graph = nx.DiGraph()
        graph.add_nodes_from(f"N{index:03d}" for index in range(1, 9))

        apply_spring_layout(graph)

        nodes = list(graph)
        for first_index, first in enumerate(nodes):
            for second in nodes[first_index + 1 :]:
                dx = (graph.nodes[second]["x"] - graph.nodes[first]["x"]) * LAYOUT_WIDTH
                dy = (graph.nodes[second]["y"] - graph.nodes[first]["y"]) * LAYOUT_HEIGHT
                self.assertGreaterEqual(math.hypot(dx, dy), MIN_NODE_CENTER_DISTANCE - 0.5)

    def test_rebuild_uses_custom_node_spacing(self) -> None:
        graph = nx.DiGraph()
        graph.add_nodes_from(f"N{index:03d}" for index in range(1, 7))

        custom_spacing = 260.0
        apply_spring_layout(graph, minimum_distance=custom_spacing)

        nodes = list(graph)
        for first_index, first in enumerate(nodes):
            for second in nodes[first_index + 1 :]:
                dx = (graph.nodes[second]["x"] - graph.nodes[first]["x"]) * LAYOUT_WIDTH
                dy = (graph.nodes[second]["y"] - graph.nodes[first]["y"]) * LAYOUT_HEIGHT
                self.assertGreaterEqual(math.hypot(dx, dy), custom_spacing - 0.5)

    def test_excel_round_trip_preserves_nodes_edges_positions_and_zero(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", label="Ремонт дорог", x=1.2, y=-0.3)
        graph.add_node("N002", label="Аварийность", x=0.8, y=0.7)
        graph.add_node("N003", label="Мобильность", x=0.5, y=0.1)
        graph.nodes["N001"]["code"] = "P02_I08"
        add_connection(
            graph,
            "N001",
            "N002",
            -0.75,
            bidirectional=True,
            reverse_weight=0.25,
            bold=True,
            reverse_bold=False,
        )
        add_connection(graph, "N003", "N001", 0.0)
        graph.graph.update(view_zoom=0.72, view_pan_x=135.0, view_pan_y=-48.0)

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "graph.xlsx"
            save_graph_to_excel(graph, path)
            restored = load_graph_from_excel(path)

        self.assertEqual(set(graph.nodes), set(restored.nodes))
        self.assertEqual(set(graph.edges), set(restored.edges))
        self.assertEqual(restored["N001"]["N002"]["weight"], -0.75)
        self.assertEqual(restored["N002"]["N001"]["weight"], 0.25)
        self.assertTrue(restored["N001"]["N002"]["bold"])
        self.assertFalse(restored["N002"]["N001"]["bold"])
        self.assertEqual(restored["N003"]["N001"]["weight"], 0.0)
        self.assertAlmostEqual(restored.nodes["N001"]["x"], 1.2)
        self.assertAlmostEqual(restored.nodes["N001"]["y"], -0.3)
        self.assertEqual(restored.nodes["N001"]["code"], "P02_I08")
        self.assertAlmostEqual(restored.graph["view_zoom"], 0.72)
        self.assertAlmostEqual(restored.graph["view_pan_x"], 135.0)
        self.assertAlmostEqual(restored.graph["view_pan_y"], -48.0)

        from_bytes = load_graph_from_excel_bytes(graph_to_excel_bytes(graph))
        self.assertEqual(from_bytes["N003"]["N001"]["weight"], 0.0)

    def test_weight_outside_range_is_rejected(self) -> None:
        graph = nx.DiGraph()
        graph.add_nodes_from(("A", "B"))
        with self.assertRaises(GraphWorkbookError):
            add_connection(graph, "A", "B", 1.01)

    def test_excel_contains_edges_with_node_labels(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", code="P02_I08", label="Причина")
        graph.add_node("N002", code="P02_I09", label="Следствие")
        graph.add_edge("N001", "N002", weight=-0.4, bold=True)

        workbook = load_workbook(BytesIO(graph_to_excel_bytes(graph)), data_only=True)
        try:
            self.assertIn("Связи по названиям", workbook.sheetnames)
            row = list(workbook["Связи по названиям"].iter_rows(min_row=2, values_only=True))[0]
            self.assertEqual(row, ("Причина", "Следствие", -0.4, True))
            self.assertIn("Связи по кодам", workbook.sheetnames)
            coded_row = list(workbook["Связи по кодам"].iter_rows(min_row=2, values_only=True))[0]
            self.assertEqual(coded_row, ("P02_I08", "P02_I09", -0.4, True))
        finally:
            workbook.close()

    def test_excel_contains_edges_with_node_descriptions(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", label="1", code="P02_06", text="Описание причины")
        graph.add_node("N002", label="2", code="P02_07", text="Описание следствия")
        graph.add_edge("N001", "N002", weight=0.6, bold=False)

        workbook = load_workbook(BytesIO(graph_to_excel_bytes(graph)), data_only=True)
        try:
            self.assertIn("Связи по описанию", workbook.sheetnames)
            rows = list(workbook["Связи по описанию"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("source", "target", "weight", "bold"))
            self.assertEqual(rows[1], ("Описание причины", "Описание следствия", 0.6, False))
        finally:
            workbook.close()

    def test_excel_contains_name_switching_sheet(self) -> None:
        graph = nx.DiGraph()
        graph.add_node(
            "N001",
            code="P02_I08",
            label="Исполнение бюджета, %",
        )

        workbook = load_workbook(BytesIO(graph_to_excel_bytes(graph)), data_only=True)
        try:
            self.assertIn("Названия", workbook.sheetnames)
            rows = list(workbook["Названия"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("id", "code", "text"))
            self.assertEqual(
                rows[1],
                ("N001", "P02_I08", "Исполнение бюджета, %"),
            )
        finally:
            workbook.close()

    def test_rename_node_preserves_id_attributes_and_edges(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", label="Старое имя", x=0.2, y=0.3, color="#4C78A8")
        graph.add_node("N002", label="Другая вершина")
        graph.add_edge("N001", "N002", weight=0.5, bold=False)

        rename_node(graph, "N001", "  Новое имя  ")

        self.assertEqual(graph.nodes["N001"]["label"], "Новое имя")
        self.assertEqual(graph.nodes["N001"]["x"], 0.2)
        self.assertTrue(graph.has_edge("N001", "N002"))

    def test_rename_node_rejects_empty_label(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", label="Имя")

        with self.assertRaises(GraphWorkbookError):
            rename_node(graph, "N001", "   ")

    def test_browser_position_updates_are_validated_and_clamped(self) -> None:
        graph = nx.DiGraph()
        graph.add_node("N001", x=0.2, y=0.3)
        graph.add_node("N002", x=0.8, y=0.7)

        changed = apply_position_updates(
            graph,
            {
                "N001": {"x": 1.2, "y": -0.1},
                "N002": {"x": "not-a-number", "y": 0.4},
                "missing": {"x": 0.5, "y": 0.5},
            },
        )

        self.assertEqual(changed, 1)
        self.assertEqual(graph.nodes["N001"]["x"], 1.2)
        self.assertEqual(graph.nodes["N001"]["y"], -0.1)
        self.assertEqual(graph.nodes["N002"]["x"], 0.8)


if __name__ == "__main__":
    unittest.main()
